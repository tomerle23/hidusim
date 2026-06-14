import openpyxl
import sys
import re
import json

sys.stdout.reconfigure(encoding='utf-8')

book_map = {
    'תהילים': 'Ps', 'תהלים': 'Ps', 'בראשית': 'Gen', 'שמות': 'Exod', 'ויקרא': 'Lev', 'במדבר': 'Num', 'דברים': 'Deut',
    'יהושע': 'Josh', 'שופטים': 'Judg', 'שמואל א': '1Sam', 'שמואל ב': '2Sam',
    'מלכים א': '1Kgs', 'מלכים ב': '2Kgs', 'ישעיהו': 'Isa', 'ירמיהו': 'Jer', 'יחזקאל': 'Ezek',
    'הושע': 'Hos', 'יואל': 'Joel', 'עמוס': 'Amos', 'עובדיה': 'Obad', 'יונה': 'Jonah', 'מיכה': 'Mic', 'נחום': 'Nah',
    'חבקוק': 'Hab', 'צפניה': 'Zeph', 'חגי': 'Hag', 'זכריה': 'Zech', 'מלאכי': 'Mal',
    'משלי': 'Prov', 'איוב': 'Job', 'שיר השירים': 'Song', 'רות': 'Ruth', 'איכה': 'Lam', 'קהלת': 'Eccl', 'אסתר': 'Esth',
    'דניאל': 'Dan', 'עזרא': 'Ezra', 'נחמיה': 'Neh', 'דברי הימים א': '1Chr', 'דברי הימים ב': '2Chr'
}

print("טוען קבצים...")
try:
    wb = openpyxl.load_workbook('1.xlsx', data_only=True)
    sheet = wb.active
except Exception as e:
    print(f"שגיאה בטעינת 1.xlsx: {e}")
    sys.exit(1)

try:
    with open('tanakh_data.js', 'r', encoding='utf-8-sig', errors='ignore') as f:
        content = f.read()
    
    json_start = content.find('{')
    json_end = content.rfind('}') + 1
    prefix = content[:json_start]
    json_str = content[json_start:json_end]
    suffix = content[json_end:]
    tanakh = json.loads(json_str)
except Exception as e:
    print(f"שגיאה בטעינת tanakh_data.js: {e}")
    sys.exit(1)

updated_count = 0
not_found = []
last_loc = None

print("משווה ומעדכן פסוקים...")
for i in range(2, sheet.max_row + 1):
    val = sheet.cell(row=i, column=1).value
    if not val:
        continue
    
    val = str(val).strip()
    
    if 'פרק' in val and 'פסוק' in val:
        last_loc = val
    else:
        if last_loc:
            parts = [p.strip() for p in re.split(r'[·:]', last_loc)]
            if len(parts) >= 3:
                book_name = parts[0]
                try:
                    chapter_num = int(re.search(r'\d+', parts[1]).group())
                    verse_num = int(re.search(r'\d+', parts[2]).group())
                    
                    eng_book = book_map.get(book_name)
                    if eng_book and eng_book in tanakh:
                        new_words = val.split()
                        current_words = tanakh[eng_book][chapter_num - 1][verse_num - 1]
                        if current_words != new_words:
                            tanakh[eng_book][chapter_num - 1][verse_num - 1] = new_words
                            updated_count += 1
                    else:
                        not_found.append((book_name, chapter_num, verse_num))
                except Exception as e:
                    print(f"שגיאה בניתוח שורה {i} מיקום {last_loc}: {e}")
            else:
                print(f"שורה {i} מיקום לא תקין: {parts}")

print(f"הסתיים. סה\"כ פסוקים שעודכנו: {updated_count}")
if not_found:
    print(f"פסוקים שלא נמצאו במאגר: {len(not_found)}")

# Save updated tanakh_data.js
try:
    corrected_js = prefix + json.dumps(tanakh, ensure_ascii=False, indent=2) + suffix
    with open('tanakh_data.js', 'w', encoding='utf-8-sig') as f:
        f.write(corrected_js)
    print("המאגר tanakh_data.js עודכן ונשמר בהצלחה.")
except Exception as e:
    print(f"שגיאה בשמירת המאגר: {e}")
